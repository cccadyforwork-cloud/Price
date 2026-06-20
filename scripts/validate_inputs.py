#!/usr/bin/env python3
import html
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "pricing_config.json"
DIMENSIONS_DIR = ROOT / "input" / "dimensions"
COMPETITOR_DIR = ROOT / "input" / "competitors"
OUTPUT_PATH = ROOT / "output" / "input_validation_report.md"


REQUIRED_DIMENSION_FIELDS = [
    "包装后长cm",
    "包装后宽cm",
    "包装后高cm",
    "包装后重量g",
]


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def pick_dimensions_file():
    files = [
        path
        for path in DIMENSIONS_DIR.glob("*.xlsx")
        if not path.name.startswith("~$")
        and "模板" not in path.name
        and "瑜伽砖" not in path.name
    ]
    if not files:
        files = [
            path
            for path in DIMENSIONS_DIR.glob("*.xlsx")
            if not path.name.startswith("~$") and "模板" not in path.name
        ]
    if not files:
        raise FileNotFoundError(f"没有找到尺寸重量表：{DIMENSIONS_DIR}")
    return max(files, key=lambda path: path.stat().st_mtime)


def strip_tags(markup):
    markup = re.sub(r"<script.*?</script>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<style.*?</style>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<[^>]+>", " ", markup)
    return re.sub(r"\s+", " ", html.unescape(markup)).strip()


def parse_pack_count(value, default=1):
    text = clean_text(value).lower()
    patterns = [
        r"(\d+)\s*(?:pcs|pc|pieces|piece|pack)",
        r"(\d+)\s*(?:支|个)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return int(match.group(1))
    return default


def load_dimensions(path):
    wb = load_workbook(path, data_only=False)
    ws = wb["产品尺寸重量输入"] if "产品尺寸重量输入" in wb.sheetnames else wb.active
    first_row = [clean_text(cell.value) for cell in ws[1]]
    rows = []

    if "SKU" in first_row:
        headers = first_row
        source_rows = ws.iter_rows(min_row=2, values_only=True)
        for row in source_rows:
            record = dict(zip(headers, row))
            sku = clean_text(record.get("SKU"))
            if not sku:
                continue
            missing = [
                field
                for field in REQUIRED_DIMENSION_FIELDS
                if clean_text(record.get(field)) == ""
            ]
            rows.append(
                {
                    "sku": sku,
                    "name": clean_text(record.get("产品名称")),
                    "variant": clean_text(record.get("颜色/规格")),
                    "qty": record.get("采购数量"),
                    "missing": missing,
                    "ready": not missing,
                }
            )
        return rows

    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if len(values) < 8 or not clean_text(values[2]):
            continue
        sku = clean_text(values[0])
        name = clean_text(values[2])
        variant = clean_text(values[3])
        size = clean_text(values[6])
        weight = clean_text(values[7])
        missing = []
        if not size:
            missing.extend(["包装后长cm", "包装后宽cm", "包装后高cm"])
        if not weight:
            missing.append("包装后重量g")
        rows.append(
            {
                "sku": sku,
                "name": name,
                "variant": variant,
                "qty": values[5],
                "missing": missing,
                "ready": not missing,
            }
        )
    return rows


def parse_competitor_html(path):
    raw = path.read_text(errors="ignore")

    title_match = re.search(
        r'id="productTitle"[^>]*>(.*?)</span>', raw, flags=re.S | re.I
    )
    title = strip_tags(title_match.group(1)) if title_match else path.stem

    price_match = re.search(
        r'<span class="a-price-whole">(\d+)'
        r'<span class="a-price-decimal">\.?</span></span>\s*'
        r'<span class="a-price-fraction">(\d+)</span>',
        raw,
    )
    price = None
    if price_match:
        price = float(f"{price_match.group(1)}.{price_match.group(2)}")

    asins = []
    for pattern in [
        r'data-asin="(B0[A-Z0-9]{8})"',
        r"/dp/(B0[A-Z0-9]{8})",
        r'"asin"\s*:\s*"(B0[A-Z0-9]{8})"',
        r"B0[A-Z0-9]{8}",
    ]:
        for match in re.finditer(pattern, raw):
            asin = match.group(1) if match.groups() else match.group(0)
            if asin not in asins:
                asins.append(asin)
        if asins:
            break

    return {
        "file": path.name,
        "title": title,
        "price": price,
        "asin": asins[0] if asins else "",
        "pack_count": parse_pack_count(title),
        "price_per_10": price / parse_pack_count(title) * 10 if price else None,
    }


def load_competitors(path):
    return [
        parse_competitor_html(file)
        for file in sorted(path.glob("*.html"))
    ]


def write_report(config, dimensions, competitors, dimensions_path):
    ready_count = sum(1 for row in dimensions if row["ready"])
    lines = [
        "# 输入数据校验报告",
        "",
        f"- 当前尺寸重量表：{dimensions_path.name}",
        "",
        "## 配置",
        "",
        f"- 站点：{config['marketplace']}",
        f"- 汇率：{config['currency_rate_rmb_to_usd']}",
        f"- 抽佣：{config['referral_fee_rate']:.0%}",
        f"- 目标最低利润率：{config['target_margin_min']:.0%}",
        f"- 体积重除数：{config['dimensional_weight_divisor']}",
        "",
        "## 尺寸重量状态",
        "",
        f"- SKU 总数：{len(dimensions)}",
        f"- 可进入定价：{ready_count}",
        f"- 待补尺寸重量：{len(dimensions) - ready_count}",
        "",
        "| SKU | 规格 | 状态 | 缺失项 |",
        "| --- | --- | --- | --- |",
    ]

    for row in dimensions:
        status = "可定价" if row["ready"] else "待补"
        missing = "、".join(row["missing"]) if row["missing"] else "-"
        lines.append(f"| {row['sku']} | {row['variant']} | {status} | {missing} |")

    lines.extend([
        "",
        "## 竞品价格",
        "",
        "| 文件 | ASIN | 页面主售价 | 包数 | 折算每10支 | 标题 |",
        "| --- | --- | ---: | ---: | ---: | --- |",
    ])

    for item in competitors:
        price = f"${item['price']:.2f}" if item["price"] is not None else "未识别"
        price_per_10 = f"${item['price_per_10']:.2f}" if item["price_per_10"] is not None else "未识别"
        title = item["title"].replace("|", "/")
        lines.append(
            f"| {item['file']} | {item['asin']} | {price} | {item['pack_count']} | {price_per_10} | {title} |"
        )

    lines.extend([
        "",
        "## 下一步",
        "",
        "如果所有 SKU 都显示“可定价”，可以进入物流费档位匹配和建议售价计算。",
        "",
    ])

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main():
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    dimensions_path = pick_dimensions_file()
    dimensions = load_dimensions(dimensions_path)
    competitors = load_competitors(COMPETITOR_DIR)
    write_report(config, dimensions, competitors, dimensions_path)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
