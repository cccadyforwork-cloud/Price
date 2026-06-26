#!/usr/bin/env python3
import json
import mimetypes
import os
import re
import socket
import subprocess
import sys
import tempfile
import uuid
from email.parser import BytesParser
from email.policy import default
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBENCH_DIR = ROOT / "workbench"
NODE_BIN = Path("/Users/cc/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node")
NODE_MODULES = Path("/Users/cc/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/node_modules")
CONFIG_PATH = ROOT / "config" / "products" / "blister_pad_bundle.json"
PRICING_CONFIG_PATH = ROOT / "config" / "pricing_config.json"
WORKBENCH_STATE_PATH = ROOT / "config" / "workbench_state.json"
OUTPUT_DIR = ROOT / "output"
INPUT_DIR = ROOT / "input"
UPLOAD_TARGETS = {
    "purchase_order": INPUT_DIR / "purchase_orders",
    "product_links": INPUT_DIR / "product_links",
    "dimensions": INPUT_DIR / "dimensions",
    "competitors": INPUT_DIR / "competitors",
}
ALLOWED_SUFFIXES = {
    "purchase_order": {".pdf", ".xlsx", ".xls", ".html", ".htm", ".txt", ".csv", ".png", ".jpg", ".jpeg", ".webp"},
    "product_links": {".txt", ".csv", ".tsv", ".xlsx", ".xls", ".html", ".htm"},
    "dimensions": {".xlsx", ".xls"},
    "competitors": {".html", ".htm"},
}


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path, data):
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def default_workbench_state():
    return {
        "product_links": [""],
        "manual_dimensions": [
            {
                "sku": "",
                "title": "",
                "length_cm": "",
                "width_cm": "",
                "height_cm": "",
                "weight_g": "",
            }
        ],
    }


def read_workbench_state():
    if not WORKBENCH_STATE_PATH.exists():
        return default_workbench_state()
    state = read_json(WORKBENCH_STATE_PATH)
    fallback = default_workbench_state()
    return {
        "product_links": state.get("product_links") or fallback["product_links"],
        "manual_dimensions": state.get("manual_dimensions") or fallback["manual_dimensions"],
    }


def write_workbench_state(data):
    state = default_workbench_state()
    if isinstance(data.get("product_links"), list):
        state["product_links"] = [str(item).strip() for item in data["product_links"]]
    if isinstance(data.get("manual_dimensions"), list):
        rows = []
        for item in data["manual_dimensions"]:
            if isinstance(item, dict):
                rows.append(
                    {
                        "sku": str(item.get("sku", "")).strip(),
                        "title": str(item.get("title", "")).strip(),
                        "length_cm": item.get("length_cm", ""),
                        "width_cm": item.get("width_cm", ""),
                        "height_cm": item.get("height_cm", ""),
                        "weight_g": item.get("weight_g", ""),
                    }
                )
        state["manual_dimensions"] = rows or state["manual_dimensions"]
    WORKBENCH_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    write_json(WORKBENCH_STATE_PATH, state)
    return state


def workbook_rows(path, sheet_name, max_rows=50):
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, max_rows + 1), values_only=True):
        if not any(value is not None and value != "" for value in values):
            continue
        rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})
    return rows


def read_report(path):
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")


def output_state(product_batch):
    draft = OUTPUT_DIR / f"{product_batch}_统一输入草稿.xlsx"
    result = OUTPUT_DIR / f"{product_batch}_正式定价结果.xlsx"
    report = OUTPUT_DIR / f"{product_batch}_正式定价报告.md"
    upload = OUTPUT_DIR / f"{product_batch}_上品系统导入.xlsx"
    draft_rows = workbook_rows(draft, "统一输入草稿", 20)
    result_rows = workbook_rows(result, "建议售价", 20)
    draft_by_sku = {row.get("SKU"): row for row in draft_rows if row.get("SKU")}
    for row in result_rows:
        source = draft_by_sku.get(row.get("SKU"), {})
        for key in ["包装长cm", "包装宽cm", "包装高cm", "包装重量g"]:
            if key in source:
                row[key] = source[key]
    return {
        "draft": {
            "path": str(draft),
            "exists": draft.exists(),
            "rows": draft_rows,
            "competitors": workbook_rows(draft, "竞品解析", 20),
        },
        "result": {
            "path": str(result),
            "exists": result.exists(),
            "rows": result_rows,
            "scenarios": workbook_rows(result, "价格档位", 20),
        },
        "report": {
            "path": str(report),
            "exists": report.exists(),
            "content": read_report(report),
        },
        "upload": {
            "path": str(upload),
            "exists": upload.exists(),
            "rows": workbook_rows(upload, "上品系统导入", 20),
        },
    }


def safe_filename(name):
    cleaned = Path(name or "upload").name.replace("\x00", "").strip()
    return cleaned or f"upload-{uuid.uuid4().hex[:8]}"


def input_files():
    state = {}
    for key, directory in UPLOAD_TARGETS.items():
        directory.mkdir(parents=True, exist_ok=True)
        state[key] = [
            {
                "name": path.name,
                "path": str(path),
                "size": path.stat().st_size,
            }
            for path in sorted(directory.iterdir(), key=lambda item: item.stat().st_mtime, reverse=True)
            if path.is_file() and not path.name.startswith(("~$", ".~", ".DS_Store"))
        ]
    return state


def parse_multipart(headers, body):
    content_type = headers.get("Content-Type", "")
    if "multipart/form-data" not in content_type:
        raise ValueError("请使用表单上传文件")
    parser_headers = (
        f"Content-Type: {content_type}\r\n"
        f"MIME-Version: 1.0\r\n\r\n"
    ).encode("utf-8")
    message = BytesParser(policy=default).parsebytes(parser_headers + body)
    fields = {}
    files = []
    for part in message.iter_parts():
        disposition = part.get_content_disposition()
        if disposition != "form-data":
            continue
        name = part.get_param("name", header="content-disposition")
        filename = part.get_filename()
        payload = part.get_payload(decode=True) or b""
        if filename:
            files.append({"field": name, "filename": filename, "payload": payload})
        elif name:
            fields[name] = payload.decode(part.get_content_charset() or "utf-8", errors="replace")
    return fields, files


def save_uploads(headers, body):
    fields, files = parse_multipart(headers, body)
    file_type = fields.get("type", "")
    if file_type not in UPLOAD_TARGETS:
        raise ValueError("未知资料类型")
    target_dir = UPLOAD_TARGETS[file_type]
    target_dir.mkdir(parents=True, exist_ok=True)
    allowed = ALLOWED_SUFFIXES[file_type]
    saved = []
    for item in files:
        original_name = safe_filename(item["filename"])
        suffix = Path(original_name).suffix.lower()
        if suffix not in allowed:
            raise ValueError(f"{original_name} 的格式不支持")
        target = target_dir / original_name
        if target.exists():
            target = target_dir / f"{target.stem}_{uuid.uuid4().hex[:6]}{target.suffix}"
        target.write_bytes(item["payload"])
        saved.append({"name": target.name, "path": str(target), "size": target.stat().st_size})
    if not saved:
        raise ValueError("没有收到文件")
    return saved


def run_image_ocr(headers, body):
    fields, files = parse_multipart(headers, body)
    if not files:
        raise ValueError("没有收到图片")
    item = files[0]
    original_name = safe_filename(item["filename"])
    suffix = Path(original_name).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg", ".webp"}:
        raise ValueError(f"{original_name} 不是支持的图片格式")
    with tempfile.NamedTemporaryFile(prefix="price-ocr-", suffix=suffix, delete=False) as tmp:
        tmp.write(item["payload"])
        tmp_path = Path(tmp.name)
    try:
        node_bin = NODE_BIN if NODE_BIN.exists() else Path("node")
        env = os.environ.copy()
        if NODE_MODULES.exists():
            env["NODE_PATH"] = str(NODE_MODULES)
        completed = subprocess.run(
            [str(node_bin), str(WORKBENCH_DIR / "ocr_node.js"), str(tmp_path)],
            cwd=ROOT,
            text=True,
            capture_output=True,
            timeout=180,
            env=env,
        )
        if completed.returncode != 0:
            raise RuntimeError(completed.stderr.strip() or "OCR 识别失败")
        payload = json.loads(completed.stdout or "{}")
        return payload.get("text", "")
    finally:
        try:
            tmp_path.unlink()
        except FileNotFoundError:
            pass


def extract_links_from_file(path):
    links = []
    if path.suffix.lower() in {".xlsx", ".xls"}:
        workbook = load_workbook(path, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    links.extend(re.findall(r"https?://\S+", str(value or "")))
    else:
        text = path.read_text(encoding="utf-8", errors="ignore")
        links.extend(re.findall(r"https?://\S+", text))
    cleaned = []
    for link in links:
        link = link.strip().strip('",;，；')
        if link and link not in cleaned:
            cleaned.append(link)
    return cleaned


def merge_product_links(saved):
    if not saved:
        return
    state = read_workbench_state()
    links = [link for link in state.get("product_links", []) if clean(link)]
    for item in saved:
        for link in extract_links_from_file(Path(item["path"])):
            if link not in links:
                links.append(link)
    if not links:
        links = [""]
    state["product_links"] = links
    write_workbench_state(state)


def run_script(script):
    completed = subprocess.run(
        [sys.executable, str(ROOT / script)],
        cwd=ROOT,
        text=True,
        capture_output=True,
        timeout=120,
    )
    return {
        "ok": completed.returncode == 0,
        "returncode": completed.returncode,
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
    }


def draft_script_for_mode(mode):
    if mode == "legacy_bundle":
        return "scripts/build_blister_pad_draft.py"
    if mode == "legacy_generic":
        return "scripts/generate_input_draft.py"
    return "scripts/build_workbench_draft.py"


def api_state():
    product = read_json(CONFIG_PATH)
    pricing = read_json(PRICING_CONFIG_PATH)
    workbench = read_workbench_state()
    purchase_order = ROOT / product["purchase_order"]
    competitor_dir = ROOT / product["competitor_dir"]
    competitor_files = []
    if competitor_dir.exists():
        competitor_files = [path.name for path in sorted(competitor_dir.glob("*.html"))]
    return {
        "product": product,
        "pricing": pricing,
        "sourceFiles": {
            "purchaseOrder": {
                "path": str(purchase_order),
                "exists": purchase_order.exists(),
            },
            "competitorDir": {
                "path": str(competitor_dir),
                "exists": competitor_dir.exists(),
                "files": competitor_files,
            },
        },
        "inputFiles": input_files(),
        "workbench": workbench,
        "outputs": output_state(product["product_batch"]),
    }


class WorkbenchHandler(BaseHTTPRequestHandler):
    server_version = "PriceWorkbench/1.0"

    def log_message(self, format, *args):
        return

    def send_json(self, data, status=200):
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def send_file(self, path):
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        content = path.read_bytes()
        mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", mime_type)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def do_GET(self):
        parsed = urlparse(self.path)
        route = parsed.path
        if route == "/api/state":
            self.send_json(api_state())
            return
        if route.startswith("/files/output/"):
            name = unquote(route.removeprefix("/files/output/"))
            path = (OUTPUT_DIR / name).resolve()
            if OUTPUT_DIR.resolve() not in path.parents and path != OUTPUT_DIR.resolve():
                self.send_error(403)
                return
            self.send_file(path)
            return
        if route in {"/", "/index.html"}:
            self.send_file(WORKBENCH_DIR / "index.html")
            return
        if route in {"/styles.css", "/app.js"}:
            self.send_file(WORKBENCH_DIR / route.removeprefix("/"))
            return
        self.send_error(404)

    def do_POST(self):
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(length) if length else b"{}"
        if parsed.path == "/api/config":
            try:
                body = raw_body.decode("utf-8")
                payload = json.loads(body)
                write_json(CONFIG_PATH, payload["product"])
                self.send_json({"ok": True, "state": api_state()})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=400)
            return
        if parsed.path == "/api/upload":
            try:
                saved = save_uploads(self.headers, raw_body)
                if saved and saved[0]["path"].startswith(str(UPLOAD_TARGETS["product_links"])):
                    merge_product_links(saved)
                self.send_json({"ok": True, "saved": saved, "state": api_state()})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=400)
            return
        if parsed.path == "/api/ocr":
            try:
                text = run_image_ocr(self.headers, raw_body)
                self.send_json({"ok": True, "text": text})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=400)
            return
        if parsed.path == "/api/workbench":
            try:
                body = raw_body.decode("utf-8")
                payload = json.loads(body)
                write_workbench_state(payload.get("workbench", {}))
                self.send_json({"ok": True, "state": api_state()})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=400)
            return
        if parsed.path == "/api/generate-draft":
            try:
                body = raw_body.decode("utf-8")
                payload = json.loads(body) if body.strip() else {}
            except json.JSONDecodeError:
                payload = {}
            result = run_script(draft_script_for_mode(payload.get("mode")))
            self.send_json({"ok": result["ok"], "command": result, "state": api_state()}, status=200 if result["ok"] else 500)
            return
        if parsed.path == "/api/run-pricing":
            try:
                body = raw_body.decode("utf-8")
                payload = json.loads(body) if body.strip() else {}
            except json.JSONDecodeError:
                payload = {}
            draft = run_script(draft_script_for_mode(payload.get("mode")))
            pricing = run_script("scripts/generate_final_pricing.py") if draft["ok"] else None
            ok = draft["ok"] and pricing and pricing["ok"]
            self.send_json(
                {"ok": ok, "draft": draft, "pricing": pricing, "state": api_state()},
                status=200 if ok else 500,
            )
            return
        self.send_error(404)


def main():
    host = "127.0.0.1"
    preferred_port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    port = preferred_port
    if len(sys.argv) == 1:
        for candidate in range(preferred_port, preferred_port + 25):
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as probe:
                if probe.connect_ex((host, candidate)) != 0:
                    port = candidate
                    break
    server = ThreadingHTTPServer((host, port), WorkbenchHandler)
    print(f"工作台已启动：http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
