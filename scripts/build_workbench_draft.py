#!/usr/bin/env python3
import html
import json
import math
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
INPUT_DIR = ROOT / "input"
PRODUCT_CONFIG_PATH = ROOT / "config" / "products" / "blister_pad_bundle.json"
WORKBENCH_STATE_PATH = ROOT / "config" / "workbench_state.json"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def to_float(value, default=0.0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def read_json(path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def strip_tags(markup):
    markup = re.sub(r"<script.*?</script>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<style.*?</style>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<[^>]+>", " ", markup)
    return re.sub(r"\s+", " ", html.unescape(markup)).strip()


def parse_pack_count(text, default=1):
    value = clean(text).lower()
    for pattern in [
        r"(\d+)\s*(?:pcs|pc|pieces|piece|pack|ct|count|pairs|pair)",
        r"(\d+)\s*(?:片|支|个|双|包)",
    ]:
        match = re.search(pattern, value)
        if match:
            return int(match.group(1))
    return default


def parse_competitor_html(path):
    raw = path.read_text(errors="ignore")
    title_match = re.search(
        r'id="productTitle"[^>]*>(.*?)</span>', raw, flags=re.S | re.I
    )
    title = strip_tags(title_match.group(1)) if title_match else path.stem
    price_match = re.search(
        r'<span class="a-price-whole">(\d+)'
        r'<span class="a-price-decimal">\.?</span></span>\s*'
        r'<span class="a-price-fraction">(\d+)</span>',
        raw,
    )
    asins = []
    for pattern in [
        r'data-asin="(B0[A-Z0-9]{8})"',
        r"/dp/(B0[A-Z0-9]{8})",
        r'"asin"\s*:\s*"(B0[A-Z0-9]{8})"',
        r"B0[A-Z0-9]{8}",
    ]:
        for match in re.finditer(pattern, raw):
            asin = match.group(1) if match.groups() else match.group(0)
            if asin not in asins:
                asins.append(asin)
        if asins:
            break
    price = float(f"{price_match.group(1)}.{price_match.group(2)}") if price_match else None
    pack_count = parse_pack_count(title)
    return {
        "文件": path.name,
        "ASIN": asins[0] if asins else "",
        "页面主售价USD": price,
        "包数": pack_count,
        "单件价USD": price / pack_count if price and pack_count else None,
        "容量ml": "",
        "标题": title,
    }


def load_competitors(product):
    candidates = []
    for directory in [ROOT / product.get("competitor_dir", "")]:
        if directory.exists() and directory.is_dir():
            candidates.extend(
                path for path in sorted(directory.glob("*.html"))
                if not path.name.startswith(("~$", ".~", ".DS_Store"))
            )
    seen = set()
    competitors = []
    for path in candidates:
        if path.resolve() in seen:
            continue
        seen.add(path.resolve())
        competitors.append(parse_competitor_html(path))
    for item in product.get("manual_competitors", []):
        price = item.get("price_usd")
        pack_count = item.get("pack_count") or 1
        competitors.append(
            {
                "文件": item.get("file", ""),
                "ASIN": item.get("asin", ""),
                "页面主售价USD": price,
                "包数": pack_count,
                "单件价USD": price / pack_count if price and pack_count else None,
                "容量ml": "",
                "标题": item.get("title", ""),
            }
        )
    return competitors


def bundle_pack_count(product):
    return sum(to_float(component.get("pcs_per_bundle")) for component in product.get("components", [])) or 1


def bundle_quantity(product):
    quantities = []
    for component in product.get("components", []):
        pcs_per_bundle = max(to_float(component.get("pcs_per_bundle")), 1)
        quantities.append(math.floor(to_float(component.get("purchase_quantity_pcs")) / pcs_per_bundle))
    return min(quantities) if quantities else 1


def bundle_purchase_cost(product):
    return sum(
        to_float(component.get("purchase_unit_cost_rmb")) * to_float(component.get("pcs_per_bundle"))
        for component in product.get("components", [])
    )


def manual_rows(workbench, product):
    rows = []
    package = product.get("package", {})
    for index, item in enumerate(workbench.get("manual_dimensions", []), start=1):
        if not any(clean(item.get(key)) for key in ["sku", "title", "length_cm", "width_cm", "height_cm", "weight_g"]):
            continue
        rows.append(
            {
                "SKU": clean(item.get("sku")) or clean(product.get("sku")) or f"SKU-{index}",
                "变体/规格": clean(item.get("title")) or clean(product.get("variant")) or f"款式{index}",
                "包装长cm": to_float(item.get("length_cm"), to_float(package.get("length_cm"))),
                "包装宽cm": to_float(item.get("width_cm"), to_float(package.get("width_cm"))),
                "包装高cm": to_float(item.get("height_cm"), to_float(package.get("height_cm"))),
                "包装重量g": to_float(item.get("weight_g"), to_float(package.get("weight_g"))),
            }
        )
    if rows:
        return rows
    return [
        {
            "SKU": clean(product.get("sku")) or "SKU-1",
            "变体/规格": clean(product.get("variant")) or "默认规格",
            "包装长cm": to_float(package.get("length_cm")),
            "包装宽cm": to_float(package.get("width_cm")),
            "包装高cm": to_float(package.get("height_cm")),
            "包装重量g": to_float(package.get("weight_g")),
        }
    ]


def build_rows(product, workbench, competitors):
    pack_count = bundle_pack_count(product)
    quantity = bundle_quantity(product)
    purchase_cost = bundle_purchase_cost(product)
    competitor_files = ";".join(item["文件"] for item in competitors)
    rows = []
    for item in manual_rows(workbench, product):
        rows.append(
            {
                "产品批次": clean(product.get("product_batch")) or "工作台产品",
                "SKU": item["SKU"],
                "中文品名": clean(product.get("chinese_name")),
                "英文品名": clean(product.get("english_name")),
                "变体/规格": item["变体/规格"],
                "销售包数": pack_count,
                "采购数量": quantity,
                "采购单价RMB/包": purchase_cost,
                "订单运费RMB": to_float(product.get("order_shipping_rmb")),
                "订单优惠RMB": to_float(product.get("order_discount_rmb")),
                "包装长cm": item["包装长cm"],
                "包装宽cm": item["包装宽cm"],
                "包装高cm": item["包装高cm"],
                "包装重量g": item["包装重量g"],
                "头程运费USD/包": to_float(product.get("first_leg_shipping_usd_per_bundle"), 0.3),
                "目标最低利润率": to_float(product.get("target_margin_min"), 0.25),
                "目标价格优势USD/每对比单位": to_float(product.get("target_price_gap_usd_per_comparison_unit"), 0.39),
                "对比单位数量": to_float(product.get("comparison_unit_quantity"), pack_count),
                "手动定价USD": product.get("manual_price_usd", ""),
                "竞品文件或链接": competitor_files,
                "备注": "工作台手动输入；确认后生成定价分析表。",
            }
        )
    return rows


def style_workbook(workbook):
    thin = Side(style="thin", color="D9D9D9")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 16


def write_workbook(product, workbench, rows, competitors):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统一输入草稿"
    headers = [
        "产品批次", "SKU", "中文品名", "英文品名", "变体/规格", "销售包数", "采购数量",
        "采购单价RMB/包", "订单运费RMB", "订单优惠RMB", "包装长cm", "包装宽cm",
        "包装高cm", "包装重量g", "头程运费USD/包", "目标最低利润率",
        "目标价格优势USD/每对比单位", "对比单位数量", "手动定价USD", "竞品文件或链接", "备注",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    competitor_sheet = workbook.create_sheet("竞品解析")
    competitor_headers = ["文件", "ASIN", "页面主售价USD", "包数", "单件价USD", "容量ml", "标题"]
    competitor_sheet.append(competitor_headers)
    for item in competitors:
        competitor_sheet.append([item.get(header, "") for header in competitor_headers])

    source_sheet = workbook.create_sheet("来源与待确认")
    source_sheet.append(["项目", "内容"])
    source_sheet.append(["数据来源", "工作台手动尺寸重量、采购单、竞品HTML和产品链接。"])
    source_sheet.append(["产品链接", "\n".join(link for link in workbench.get("product_links", []) if clean(link))])
    source_sheet.append(["规则", "读取项目内置 config/pricing_config.json 和 input/rules/。"])
    source_sheet.append(["下一步", "确认草稿后运行正式定价，输出定价分析表和Excel。"])

    style_workbook(workbook)
    sheet.column_dimensions["T"].width = 56
    sheet.column_dimensions["U"].width = 72
    competitor_sheet.column_dimensions["A"].width = 42
    competitor_sheet.column_dimensions["G"].width = 80
    source_sheet.column_dimensions["B"].width = 100

    product_batch = clean(product.get("product_batch")) or "工作台产品"
    output_path = OUTPUT_DIR / f"{product_batch}_统一输入草稿.xlsx"
    workbook.save(output_path)
    return output_path


def main():
    product = read_json(PRODUCT_CONFIG_PATH, {})
    workbench = read_json(WORKBENCH_STATE_PATH, {})
    competitors = load_competitors(product)
    rows = build_rows(product, workbench, competitors)
    output_path = write_workbook(product, workbench, rows, competitors)
    print(output_path)


if __name__ == "__main__":
    main()
