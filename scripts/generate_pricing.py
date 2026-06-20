#!/usr/bin/env python3
import html
import json
import math
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "pricing_config.json"
DIMENSIONS_DIR = ROOT / "input" / "dimensions"
COMPETITOR_DIR = ROOT / "input" / "competitors"


FBA_FEE_TIERS = [
    {"label": "4oz及以下且售价<=3", "max_lb": 4 / 16, "max_price": 3, "fee": 0.50},
    {"label": "4oz及以下且售价>3", "max_lb": 4 / 16, "min_price": 3, "fee": 0.88},
    {"label": "4+~8oz", "max_lb": 8 / 16, "fee": 1.77},
    {"label": "8+~12oz", "max_lb": 12 / 16, "fee": 2.60},
    {"label": "12+~16oz", "max_lb": 1.0, "fee": 3.22},
    {"label": "1+~1.25lb", "max_lb": 1.25, "fee": 3.72},
    {"label": "1.25+~1.5lb", "max_lb": 1.5, "fee": 4.42},
    {"label": "1.5+~1.75lb", "max_lb": 1.75, "fee": 5.11},
    {"label": "1.75+~2lb", "max_lb": 2.0, "fee": 5.81},
    {"label": "2+~2.5lb", "max_lb": 2.5, "fee": 6.83},
    {"label": "2.5+~3lb", "max_lb": 3.0, "fee": 8.19},
    {"label": "3+~3.5lb", "max_lb": 3.5, "fee": 9.59},
    {"label": "3.5+~4lb", "max_lb": 4.0, "fee": 11.05},
    {"label": "4lb+", "max_lb": math.inf, "fee": 12.51},
]


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def pick_dimensions_file():
    files = [
        path
        for path in DIMENSIONS_DIR.glob("*.xlsx")
        if not path.name.startswith("~$")
        and "模板" not in path.name
        and "瑜伽砖" not in path.name
    ]
    if not files:
        files = [
            path
            for path in DIMENSIONS_DIR.glob("*.xlsx")
            if not path.name.startswith("~$") and "模板" not in path.name
        ]
    if not files:
        raise FileNotFoundError(f"没有找到尺寸重量表：{DIMENSIONS_DIR}")
    return max(files, key=lambda path: path.stat().st_mtime)


def output_paths(dimensions_path):
    product_name = dimensions_path.stem.replace("尺寸重量输入", "").replace("尺寸", "")
    product_name = product_name.strip("_ -") or dimensions_path.stem
    return (
        ROOT / "output" / f"{product_name}_自动定价结果.xlsx",
        ROOT / "output" / f"{product_name}_自动定价报告.md",
        product_name,
    )


def to_float(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text.startswith("="):
        return default
    try:
        return float(text)
    except ValueError:
        return default


def parse_size_cm(value):
    text = clean_text(value).lower().replace("×", "*").replace("x", "*")
    parts = re.findall(r"\d+(?:\.\d+)?", text)
    if len(parts) < 3:
        return None
    return [float(parts[0]), float(parts[1]), float(parts[2])]


def parse_weight_g(value):
    text = clean_text(value).lower()
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    weight = float(match.group(0))
    if "kg" in text:
        return weight * 1000
    return weight


def parse_pack_count(value, default=1):
    text = clean_text(value).lower()
    patterns = [
        r"(\d+)\s*(?:pcs|pc|pieces|piece|pack)",
        r"(\d+)\s*(?:支|个)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return int(match.group(1))
    return default


def strip_tags(markup):
    markup = re.sub(r"<script.*?</script>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<style.*?</style>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<[^>]+>", " ", markup)
    return re.sub(r"\s+", " ", html.unescape(markup)).strip()


def parse_competitor_html(path):
    raw = path.read_text(errors="ignore")
    title_match = re.search(
        r'id="productTitle"[^>]*>(.*?)</span>', raw, flags=re.S | re.I
    )
    price_match = re.search(
        r'<span class="a-price-whole">(\d+)'
        r'<span class="a-price-decimal">\.?</span></span>\s*'
        r'<span class="a-price-fraction">(\d+)</span>',
        raw,
    )
    asins = []
    for pattern in [
        r'data-asin="(B0[A-Z0-9]{8})"',
        r"/dp/(B0[A-Z0-9]{8})",
        r'"asin"\s*:\s*"(B0[A-Z0-9]{8})"',
        r"B0[A-Z0-9]{8}",
    ]:
        for match in re.finditer(pattern, raw):
            asin = match.group(1) if match.groups() else match.group(0)
            if asin not in asins:
                asins.append(asin)
        if asins:
            break
    title = strip_tags(title_match.group(1)) if title_match else path.stem
    return {
        "file": path.name,
        "asin": asins[0] if asins else "",
        "title": title,
        "price": float(f"{price_match.group(1)}.{price_match.group(2)}")
        if price_match
        else None,
        "pack_count": parse_pack_count(title),
    }


def load_competitors():
    competitors = [
        parse_competitor_html(path)
        for path in sorted(COMPETITOR_DIR.glob("*.html"))
    ]
    for item in competitors:
        item["price_per_10"] = (
            item["price"] / item["pack_count"] * 10
            if item["price"] is not None and item["pack_count"]
            else None
        )
    prices = [item["price_per_10"] for item in competitors if item["price_per_10"] is not None]
    return competitors, min(prices) if prices else None


def load_skus(dimensions_path):
    wb = load_workbook(dimensions_path, data_only=False)
    ws = wb["产品尺寸重量输入"] if "产品尺寸重量输入" in wb.sheetnames else wb.active
    first_row = [clean_text(cell.value) for cell in ws[1]]
    rows = []
    if "SKU" in first_row:
        headers = first_row
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            sku = clean_text(record.get("SKU"))
            if not sku:
                continue
            rows.append(record)
        return rows

    generic_rows = []
    for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(row)
        if len(values) < 8 or not clean_text(values[2]):
            continue
        size = parse_size_cm(values[6])
        if size is None:
            continue
        generic_rows.append(
            {
                "SKU": f"FLOWER-CARD-{clean_text(values[0]) or index}",
                "产品名称": clean_text(values[2]),
                "颜色/规格": clean_text(values[3]),
                "采购数量": to_float(values[5]),
                "采购单价RMB": to_float(values[4]),
                "国内运费总额RMB": 0,
                "包装后长cm": size[0],
                "包装后宽cm": size[1],
                "包装后高cm": size[2],
                "包装后重量g": parse_weight_g(values[7]),
                "销售套装数量": parse_pack_count(values[3], default=10),
            }
        )
    total_qty = sum(to_float(row["采购数量"]) for row in generic_rows)
    net_order_extra_rmb = 2.0 if "花束卡片夹" in dimensions_path.stem else 0.0
    for row in generic_rows:
        row["国内运费总额RMB"] = net_order_extra_rmb
    return generic_rows


def fba_fee_for(weight_lb, price):
    for tier in FBA_FEE_TIERS:
        if weight_lb > tier["max_lb"]:
            continue
        if "max_price" in tier and price > tier["max_price"]:
            continue
        if "min_price" in tier and price <= tier["min_price"]:
            continue
        return tier["fee"], tier["label"]
    return FBA_FEE_TIERS[-1]["fee"], FBA_FEE_TIERS[-1]["label"]


def unit_domestic_freight(rows):
    total_qty = sum(to_float(row.get("采购数量")) for row in rows)
    totals = [
        to_float(row.get("国内运费总额RMB"))
        for row in rows
        if to_float(row.get("国内运费总额RMB")) > 0
    ]
    freight_total = totals[0] if totals else 0
    return freight_total / total_qty if total_qty else 0


def metrics(price, sku, config, domestic_freight_rmb):
    pack_count = max(to_float(sku.get("销售套装数量"), 1), 1)
    length_in = to_float(sku.get("包装后长cm")) / 2.54
    width_in = to_float(sku.get("包装后宽cm")) / 2.54
    height_in = to_float(sku.get("包装后高cm")) / 2.54
    actual_lb = to_float(sku.get("包装后重量g")) / 453.59237
    dimensional_lb = length_in * width_in * height_in / config["dimensional_weight_divisor"]
    billable_lb = max(actual_lb, dimensional_lb)
    fba_fee, fba_tier = fba_fee_for(billable_lb, price)

    unit_cost_rmb = to_float(sku.get("采购单价RMB")) + domestic_freight_rmb
    unit_cost_usd = unit_cost_rmb / config["currency_rate_rmb_to_usd"]
    compared_price_per_10 = price / pack_count * 10
    referral_fee = price * config["referral_fee_rate"]
    sold_profit = (
        price
        - referral_fee
        - fba_fee
        - config["first_leg_shipping_usd"]
        - unit_cost_usd
    )
    return_loss = (
        unit_cost_usd
        + referral_fee * 0.2
        + config["first_leg_shipping_usd"]
        + fba_fee
        + config["disposal_fee_usd"]
    )
    expected_profit = (
        (1 - config["return_rate"]) * sold_profit
        - config["return_rate"] * return_loss
    )
    margin = expected_profit / price if price else 0
    return {
        "length_in": length_in,
        "width_in": width_in,
        "height_in": height_in,
        "actual_lb": actual_lb,
        "dimensional_lb": dimensional_lb,
        "billable_lb": billable_lb,
        "fba_fee": fba_fee,
        "fba_tier": fba_tier,
        "unit_cost_rmb": unit_cost_rmb,
        "unit_cost_usd": unit_cost_usd,
        "pack_count": pack_count,
        "compared_price_per_10": compared_price_per_10,
        "referral_fee": referral_fee,
        "sold_profit": sold_profit,
        "return_loss": return_loss,
        "expected_profit": expected_profit,
        "margin": margin,
    }


def sale_price_from_per_10(price_per_10, sku):
    pack_count = max(to_float(sku.get("销售套装数量"), 1), 1)
    return round(price_per_10 / 10 * pack_count, 2)


def price_needed_for_margin(sku, config, domestic_freight_rmb, target_margin):
    best = None
    price = 2.99
    while price <= 39.99:
        item = metrics(round(price, 2), sku, config, domestic_freight_rmb)
        if item["margin"] >= target_margin:
            best = round(price, 2)
            break
        price += 0.01
    return best


def build_results(config, rows, competitor_min_per_10):
    domestic_freight_rmb = unit_domestic_freight(rows)
    target_competitive_per_10 = (
        round(competitor_min_per_10 - config["target_price_gap_usd"], 2)
        if competitor_min_per_10 is not None
        else None
    )
    results = []
    for row in rows:
        target_margin_price = price_needed_for_margin(
            row, config, domestic_freight_rmb, config["target_margin_min"]
        )
        prices = [2.99, 3.49, 3.99, 4.49]
        target_competitive_price = None
        competitor_min_price = None
        if competitor_min_per_10 is not None:
            target_competitive_price = sale_price_from_per_10(target_competitive_per_10, row)
            competitor_min_price = sale_price_from_per_10(competitor_min_per_10, row)
            prices.extend([target_competitive_price, competitor_min_price])
        if target_margin_price is not None:
            prices.append(target_margin_price)
        prices = sorted({round(price, 2) for price in prices if price and price > 0})

        scenarios = []
        for price in prices:
            item = metrics(price, row, config, domestic_freight_rmb)
            label = "测算价"
            if target_competitive_price is not None and price == target_competitive_price:
                label = "竞品优势价(按每10支)"
            if competitor_min_price is not None and price == competitor_min_price:
                label = "竞品最低价(按每10支)"
            if target_margin_price is not None and price == target_margin_price:
                label = f"目标{config['target_margin_min']:.0%}利润率价"
            scenarios.append({"price": price, "label": label, **item})

        recommended = None
        if target_competitive_price is not None:
            competitive = metrics(target_competitive_price, row, config, domestic_freight_rmb)
            if competitive["margin"] >= config["target_margin_min"]:
                recommended = target_competitive_price
        if recommended is None:
            recommended = target_margin_price

        recommended_metrics = (
            metrics(recommended, row, config, domestic_freight_rmb)
            if recommended is not None
            else None
        )
        conclusion = "可低价竞争"
        if recommended is None:
            conclusion = "无法达到目标利润率"
        elif competitor_min_price is not None and recommended >= competitor_min_price:
            conclusion = "价格优势不足"
        elif recommended_metrics and recommended_metrics["expected_profit"] < 0:
            conclusion = "亏损风险"

        results.append(
            {
                "sku": clean_text(row.get("SKU")),
                "name": clean_text(row.get("产品名称")),
                "variant": clean_text(row.get("颜色/规格")),
                "recommended_price": recommended,
                "competitor_min_price": competitor_min_price,
                "competitor_min_per_10": competitor_min_per_10,
                "recommended_metrics": recommended_metrics,
                "conclusion": conclusion,
                "scenarios": scenarios,
            }
        )
    return results


def write_xlsx(results, competitors, config, competitor_min_per_10, output_xlsx):
    wb = Workbook()
    ws = wb.active
    ws.title = "建议售价"
    headers = [
        "SKU", "规格", "建议整包售价", "你的每10支价格", "结论", "计费重量lb", "FBA档位", "FBA费",
        "期望利润", "期望利润率", "竞品每10支最低价", "相对竞品差价", "风险提示"
    ]
    ws.append(headers)

    for result in results:
        m = result["recommended_metrics"]
        if m is None:
            ws.append([result["sku"], result["variant"], "", result["conclusion"]])
            continue
        your_per_10 = m["compared_price_per_10"]
        gap = your_per_10 - competitor_min_per_10 if competitor_min_per_10 else ""
        risk = (
            "体积重显著高于实重，请复核包装尺寸"
            if m["dimensional_lb"] > m["actual_lb"] * 3
            else ""
        )
        ws.append([
            result["sku"],
            result["variant"],
            result["recommended_price"],
            your_per_10,
            result["conclusion"],
            m["billable_lb"],
            m["fba_tier"],
            m["fba_fee"],
            m["expected_profit"],
            m["margin"],
            competitor_min_per_10,
            gap,
            risk,
        ])

    detail = wb.create_sheet("价格档位")
    detail.append([
        "SKU", "规格", "档位", "售价", "计费重量lb", "实重lb", "体积重lb",
        "FBA档位", "FBA费", "售出利润", "退货亏损", "期望利润", "期望利润率"
    ])
    for result in results:
        for item in result["scenarios"]:
            detail.append([
                result["sku"], result["variant"], item["label"], item["price"],
                item["billable_lb"], item["actual_lb"], item["dimensional_lb"],
                item["fba_tier"], item["fba_fee"], item["sold_profit"],
                item["return_loss"], item["expected_profit"], item["margin"],
            ])

    comp = wb.create_sheet("竞品")
    comp.append(["文件", "ASIN", "页面主售价", "包数", "折算每10支", "标题"])
    for item in competitors:
        comp.append([
            item["file"],
            item["asin"],
            item["price"],
            item["pack_count"],
            item["price_per_10"],
            item["title"],
        ])

    style_workbook(wb)
    for sheet in [ws, detail]:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"
        for cell in sheet["I"]:
            if isinstance(cell.value, float):
                cell.number_format = "0.0%"
    for cell in detail["M"]:
        if isinstance(cell.value, float):
            cell.number_format = "0.0%"
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def style_workbook(wb):
    thin = Side(style="thin", color="D9D9D9")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = "A2"
    wb["建议售价"].column_dimensions["L"].width = 34
    wb["竞品"].column_dimensions["D"].width = 80


def write_markdown(results, competitors, config, competitor_min_per_10, output_md, product_name):
    first_metrics = next(
        (
            result["recommended_metrics"]
            for result in results
            if result["recommended_metrics"] is not None
        ),
        None,
    )
    size_note = ""
    if first_metrics is not None:
        size_note = (
            f"当前包装尺寸换算后体积重约 {first_metrics['dimensional_lb']:.2f} lb，"
            f"实重约 {first_metrics['actual_lb']:.2f} lb，"
            f"系统按较大的 {first_metrics['billable_lb']:.2f} lb 作为计费重量。"
        )

    lines = [
        f"# {product_name}自动定价报告",
        "",
        f"- 竞品最低每10支价：${competitor_min_per_10:.2f}" if competitor_min_per_10 else "- 竞品最低每10支价：未识别",
        f"- 目标价格优势：每10支低于竞品 ${config['target_price_gap_usd']:.2f}",
        f"- 目标最低利润率：{config['target_margin_min']:.0%}",
        "",
        "## 建议售价",
        "",
        "| SKU | 规格 | 建议整包售价 | 你的每10支价格 | 结论 | 计费重量 | FBA费 | 期望利润率 | 风险 |",
        "| --- | --- | ---: | ---: | --- | ---: | ---: | ---: | --- |",
    ]
    for result in results:
        m = result["recommended_metrics"]
        if m is None:
            lines.append(
                f"| {result['sku']} | {result['variant']} | - | - | {result['conclusion']} | - | - | - | - |"
            )
            continue
        risk = "体积重显著高于实重，复核包装尺寸" if m["dimensional_lb"] > m["actual_lb"] * 3 else "-"
        lines.append(
            f"| {result['sku']} | {result['variant']} | ${result['recommended_price']:.2f} | "
            f"${m['compared_price_per_10']:.2f} | {result['conclusion']} | "
            f"{m['billable_lb']:.2f} lb | ${m['fba_fee']:.2f} | "
            f"{m['margin']:.1%} | {risk} |"
        )

    lines.extend([
        "",
        "## 重要判断",
        "",
        "本报告按“每10支价格”对齐竞品包数，避免把你的10支装和竞品30/40/50支装整包价直接比较。",
        size_note,
        "如果体积重显著高于实重，说明尺寸仍然是主要成本来源，需要复核是否为单个销售单位的最终包装尺寸。",
        "如果这是外箱尺寸、多个产品打包尺寸，或者估算尺寸，需要改成单个买家收到的包装尺寸后重新运行。",
        "",
        "## 竞品",
        "",
    ])
    for item in competitors:
        price = f"${item['price']:.2f}" if item["price"] is not None else "未识别"
        per_10 = f"${item['price_per_10']:.2f}" if item["price_per_10"] is not None else "未识别"
        lines.append(
            f"- {price} / {item['pack_count']}支 / 每10支{per_10} / {item['asin']} / {item['title']}"
        )
    output_md.write_text("\n".join(lines), encoding="utf-8")


def main():
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    dimensions_path = pick_dimensions_file()
    output_xlsx, output_md, product_name = output_paths(dimensions_path)
    rows = load_skus(dimensions_path)
    competitors, competitor_min_per_10 = load_competitors()
    results = build_results(config, rows, competitor_min_per_10)
    write_xlsx(results, competitors, config, competitor_min_per_10, output_xlsx)
    write_markdown(results, competitors, config, competitor_min_per_10, output_md, product_name)
    print(output_xlsx)
    print(output_md)


if __name__ == "__main__":
    main()
