#!/usr/bin/env python3
import html
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"
PURCHASE_DIR = INPUT_DIR / "purchase_orders"
DIMENSIONS_DIR = INPUT_DIR / "dimensions"
COMPETITOR_DIR = INPUT_DIR / "competitors"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def to_float(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else default


def strip_tags(markup):
    markup = re.sub(r"<script.*?</script>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<style.*?</style>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<[^>]+>", " ", markup)
    return re.sub(r"\s+", " ", html.unescape(markup)).strip()


def parse_pack_count(text, default=1):
    value = clean(text).lower()
    for pattern in [
        r"(\d+)\s*(?:pairs|pair)",
        r"(\d+)\s*(?:pcs|pc|pieces|piece|pack)",
        r"(\d+)\s*(?:支|个|双)",
    ]:
        match = re.search(pattern, value)
        if match:
            return int(match.group(1))
    return default


def parse_capacity_ml(text):
    match = re.search(r"(\d+)\s*ml", clean(text).lower())
    return int(match.group(1)) if match else ""


def parse_weight_g(text):
    value = clean(text).lower()
    match = re.search(r"\d+(?:\.\d+)?", value)
    if not match:
        return ""
    weight = float(match.group(0))
    return weight * 1000 if "kg" in value else weight


def parse_size_triplets(text):
    value = clean(text).lower().replace("×", "*").replace("x", "*")
    return [
        tuple(float(part) for part in match)
        for match in re.findall(
            r"(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)",
            value,
        )
    ]


def infer_package_size(size_text):
    parts = parse_size_triplets(size_text)
    if not parts:
        return "", "", "", "未识别尺寸，请补充最终包装尺寸"
    if len(parts) == 1:
        return (*parts[0], "尺寸来自尺寸重量表")
    length = max(part[0] for part in parts)
    width = max(part[1] for part in parts)
    height = sum(part[2] for part in parts)
    return length, width, height, "由多个部件尺寸推断，请确认最终包装尺寸"


def pick_latest_file(directory, suffixes, exclude_keywords=()):
    files = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and not path.name.startswith(("~$", ".~", ".DS_Store"))
        and path.suffix.lower() in suffixes
        and not any(keyword in path.name for keyword in exclude_keywords)
    ]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def product_name_from_path(path):
    name = path.stem
    for token in ["尺寸重量", "尺寸", "采购单", "订单详情"]:
        name = name.replace(token, "")
    if name == "玻璃壶":
        name = "玻璃喷壶"
    return name.strip("_ -") or path.stem


def read_dimension_rows(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["产品输入"] if "产品输入" in workbook.sheetnames else workbook.active
    first_row = [clean(cell.value) for cell in sheet[1]]

    if "SKU" in first_row and "包装长cm" in first_row:
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            record = dict(zip(first_row, values))
            if clean(record.get("SKU")):
                rows.append(record)
        return rows

    rows = []
    for index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = list(values)
        if len(values) < 8 or not clean(values[2]):
            continue
        length, width, height, size_note = infer_package_size(values[6])
        rows.append(
            {
                "source_index": clean(values[0]) or str(index),
                "中文品名": clean(values[2]),
                "变体/规格": clean(values[3]),
                "销售包数": parse_pack_count(values[3]),
                "采购数量": to_float(values[5]),
                "采购单价RMB/包": to_float(values[4]),
                "包装长cm": length,
                "包装宽cm": width,
                "包装高cm": height,
                "包装重量g": parse_weight_g(values[7]),
                "size_note": size_note,
                "raw_size": clean(values[6]),
            }
        )
    return rows


def parse_competitor_html(path):
    raw = path.read_text(errors="ignore")
    title_match = re.search(
        r'id="productTitle"[^>]*>(.*?)</span>', raw, flags=re.S | re.I
    )
    title = strip_tags(title_match.group(1)) if title_match else path.stem
    price_match = re.search(
        r'<span class="a-price-whole">(\d+)'
        r'<span class="a-price-decimal">\.?</span></span>\s*'
        r'<span class="a-price-fraction">(\d+)</span>',
        raw,
    )
    asins = []
    for pattern in [
        r'data-asin="(B0[A-Z0-9]{8})"',
        r"/dp/(B0[A-Z0-9]{8})",
        r'"asin"\s*:\s*"(B0[A-Z0-9]{8})"',
        r"B0[A-Z0-9]{8}",
    ]:
        for match in re.finditer(pattern, raw):
            asin = match.group(1) if match.groups() else match.group(0)
            if asin not in asins:
                asins.append(asin)
        if asins:
            break
    price = float(f"{price_match.group(1)}.{price_match.group(2)}") if price_match else None
    pack_count = parse_pack_count(title)
    return {
        "文件": path.name,
        "ASIN": asins[0] if asins else "",
        "页面主售价USD": price,
        "包数": pack_count,
        "单件价USD": price / pack_count if price and pack_count else None,
        "容量ml": parse_capacity_ml(title),
        "标题": title,
    }


def read_competitors(product_name):
    candidates = [
        path
        for path in COMPETITOR_DIR.glob("*.html")
        if product_name not in {"", "标准输入模板"}
    ]
    if "玻璃" in product_name:
        candidates = [
            path for path in candidates
            if any(token in path.name for token in ["Glass", "OFFIDIX", "Ebristar", "Mister"])
        ]
    elif "花束" in product_name or "卡片" in product_name:
        candidates = [
            path for path in candidates
            if any(token.lower() in path.name.lower() for token in ["floral", "card_holder", "flower"])
        ]
    if not candidates:
        candidates = sorted(COMPETITOR_DIR.glob("*.html"), key=lambda path: path.stat().st_mtime)[-3:]
    return [parse_competitor_html(path) for path in sorted(candidates)]


def sku_for(product_name, row):
    variant = clean(row.get("变体/规格"))
    if "玻璃" in product_name:
        colors = [
            ("款式1", "Purple", "紫色 200ml"),
            ("款式2", "Gray", "灰色 200ml"),
            ("款式3", "Brown", "棕色 200ml"),
            ("款式4", "Yellow", "黄色 200ml"),
        ]
        for key, color, label in colors:
            if key in variant:
                row["变体/规格"] = label
                return f"GlassMister-{color}-200ml"
        return f"GlassMister-Style{row.get('source_index')}-200ml"
    return f"{product_name}-{row.get('source_index')}"


def build_draft(product_name, dimension_rows, competitors, purchase_path):
    total_qty = sum(to_float(row.get("采购数量")) for row in dimension_rows)
    order_shipping = (
        10.80 if "玻璃" in product_name
        else 5.00 if "花束" in product_name
        else 3.30 if "园艺手套" in product_name
        else 0
    )
    order_discount = -1.00 if "玻璃" in product_name else (-3.00 if "花束" in product_name else 0)

    output_rows = []
    for row in dimension_rows:
        output_rows.append(
            {
                "产品批次": product_name,
                "SKU": clean(row.get("SKU")) or sku_for(product_name, row),
                "中文品名": clean(row.get("中文品名")) or product_name,
                "英文品名": clean(row.get("英文品名")) or "",
                "变体/规格": clean(row.get("变体/规格")),
                "销售包数": to_float(row.get("销售包数"), 1),
                "采购数量": to_float(row.get("采购数量")),
                "采购单价RMB/包": to_float(row.get("采购单价RMB/包")),
                "订单运费RMB": order_shipping,
                "订单优惠RMB": order_discount,
                "包装长cm": row.get("包装长cm"),
                "包装宽cm": row.get("包装宽cm"),
                "包装高cm": row.get("包装高cm"),
                "包装重量g": row.get("包装重量g"),
                "头程运费USD/包": 0.30,
                "目标最低利润率": 0.25,
                "目标价格优势USD/每对比单位": 0.39,
                "对比单位数量": to_float(row.get("销售包数"), 1),
                "手动定价USD": "",
                "竞品文件或链接": ";".join(item["文件"] for item in competitors),
                "备注": f"尺寸来源：{row.get('raw_size', '')}；{row.get('size_note', '')}",
            }
        )
    source_notes = [
        ["项目", "内容"],
        ["采购单", f"{purchase_path.name if purchase_path else '未识别'}；运费/优惠如未自动识别，请检查草稿。"],
        ["尺寸重量", "已从最新尺寸重量表生成草稿。"],
        ["尺寸风险", "如果备注提示由部件尺寸推断，请确认最终包装尺寸。"],
        ["竞品", f"已解析 {len(competitors)} 个竞品网页。"],
        ["下一步", "请检查 SKU、最终包装尺寸、包装重量、头程运费和手动定价。确认后运行 generate_final_pricing.py。"],
    ]
    return output_rows, source_notes


def write_workbook(product_name, rows, competitors, source_notes):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{product_name}_统一输入草稿.xlsx"
    if output_path.exists():
        print(f"{output_path} 已存在，未覆盖。需要重新生成时请先改名或删除旧草稿。")
        return output_path
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统一输入草稿"
    headers = [
        "产品批次", "SKU", "中文品名", "英文品名", "变体/规格", "销售包数", "采购数量",
        "采购单价RMB/包", "订单运费RMB", "订单优惠RMB", "包装长cm", "包装宽cm",
        "包装高cm", "包装重量g", "头程运费USD/包", "目标最低利润率",
        "目标价格优势USD/每对比单位", "对比单位数量", "手动定价USD", "竞品文件或链接", "备注",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    competitor_sheet = workbook.create_sheet("竞品解析")
    competitor_headers = ["文件", "ASIN", "页面主售价USD", "包数", "单件价USD", "容量ml", "标题"]
    competitor_sheet.append(competitor_headers)
    for item in competitors:
        competitor_sheet.append([item.get(header, "") for header in competitor_headers])

    source_sheet = workbook.create_sheet("来源与待确认")
    for row in source_notes:
        source_sheet.append(row)

    style_workbook(workbook)
    sheet.column_dimensions["T"].width = 48
    sheet.column_dimensions["U"].width = 56
    competitor_sheet.column_dimensions["A"].width = 56
    competitor_sheet.column_dimensions["G"].width = 80
    source_sheet.column_dimensions["A"].width = 16
    source_sheet.column_dimensions["B"].width = 100
    workbook.save(output_path)
    return output_path


def style_workbook(workbook):
    thin = Side(style="thin", color="D9D9D9")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = max(
                sheet.column_dimensions[get_column_letter(col)].width or 13,
                13,
            )


def main():
    dimension_path = pick_latest_file(
        DIMENSIONS_DIR,
        {".xlsx"},
        exclude_keywords=("模板", "瑜伽砖"),
    )
    if not dimension_path:
        raise FileNotFoundError("没有找到可用的尺寸重量表")
    purchase_path = pick_latest_file(PURCHASE_DIR, {".pdf", ".xlsx"})
    product_name = product_name_from_path(dimension_path)
    dimension_rows = read_dimension_rows(dimension_path)
    competitors = read_competitors(product_name)
    rows, source_notes = build_draft(product_name, dimension_rows, competitors, purchase_path)
    output_path = write_workbook(product_name, rows, competitors, source_notes)
    print(output_path)


if __name__ == "__main__":
    main()
