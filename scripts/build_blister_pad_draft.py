#!/usr/bin/env python3
import html
import json
import math
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
CONFIG_PATH = ROOT / "config" / "products" / "blister_pad_bundle.json"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def strip_tags(markup):
    markup = re.sub(r"<script.*?</script>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<style.*?</style>", " ", markup, flags=re.S | re.I)
    markup = re.sub(r"<[^>]+>", " ", markup)
    return re.sub(r"\s+", " ", html.unescape(markup)).strip()


def parse_pack_count(text, default=1):
    value = clean(text).lower()
    for pattern in [
        r"(\d+)\s*(?:pcs|pc|pieces|piece|pack|ct|count)",
        r"(\d+)\s*(?:片|支|个|包)",
    ]:
        match = re.search(pattern, value)
        if match:
            return int(match.group(1))
    return default


def parse_competitor_html(path):
    raw = path.read_text(errors="ignore")
    title_match = re.search(
        r'id="productTitle"[^>]*>(.*?)</span>', raw, flags=re.S | re.I
    )
    title = strip_tags(title_match.group(1)) if title_match else path.stem
    price_match = re.search(
        r'<span class="a-price-whole">(\d+)'
        r'<span class="a-price-decimal">\.?</span></span>\s*'
        r'<span class="a-price-fraction">(\d+)</span>',
        raw,
    )
    asins = []
    for pattern in [
        r'data-asin="(B0[A-Z0-9]{8})"',
        r"/dp/(B0[A-Z0-9]{8})",
        r'"asin"\s*:\s*"(B0[A-Z0-9]{8})"',
        r"B0[A-Z0-9]{8}",
    ]:
        for match in re.finditer(pattern, raw):
            asin = match.group(1) if match.groups() else match.group(0)
            if asin not in asins:
                asins.append(asin)
        if asins:
            break
    price = float(f"{price_match.group(1)}.{price_match.group(2)}") if price_match else None
    pack_count = parse_pack_count(title)
    return {
        "文件": path.name,
        "ASIN": asins[0] if asins else "",
        "页面主售价USD": price,
        "包数": pack_count,
        "单件价USD": price / pack_count if price and pack_count else None,
        "容量ml": "",
        "标题": title,
    }


def load_config():
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def load_competitors(config):
    competitor_dir = ROOT / config["competitor_dir"]
    competitors = [
        parse_competitor_html(path)
        for path in sorted(competitor_dir.glob("*.html"))
        if not path.name.startswith(("~$", ".~", ".DS_Store"))
    ]
    for item in config.get("manual_competitors", []):
        price = item.get("price_usd")
        pack_count = item.get("pack_count") or 1
        competitors.append(
            {
                "文件": item.get("file", ""),
                "ASIN": item.get("asin", ""),
                "页面主售价USD": price,
                "包数": pack_count,
                "单件价USD": price / pack_count if price and pack_count else None,
                "容量ml": "",
                "标题": item.get("title", ""),
            }
        )
    return competitors


def bundle_quantity(config):
    quantities = []
    for component in config["components"]:
        pcs_per_bundle = component["pcs_per_bundle"]
        quantities.append(math.floor(component["purchase_quantity_pcs"] / pcs_per_bundle))
    return min(quantities) if quantities else 0


def bundle_purchase_cost(config):
    return sum(
        component["purchase_unit_cost_rmb"] * component["pcs_per_bundle"]
        for component in config["components"]
    )


def bundle_pack_count(config):
    return sum(component["pcs_per_bundle"] for component in config["components"])


def component_summary(config):
    return "；".join(
        (
            f"{component['name']}：{component['purchase_quantity_pcs']}片，"
            f"{component['purchase_unit_cost_rmb']:.2f}元/片，"
            f"每套用{component['pcs_per_bundle']}片"
        )
        for component in config["components"]
    )


def build_rows(config, competitors):
    competitor_files = ";".join(item["文件"] for item in competitors)
    package = config["package"]
    pack_count = bundle_pack_count(config)
    quantity = bundle_quantity(config)
    purchase_cost = bundle_purchase_cost(config)
    manual_price = config.get("manual_price_usd", "")
    row = {
        "产品批次": config["product_batch"],
        "SKU": config["sku"],
        "中文品名": config["chinese_name"],
        "英文品名": config["english_name"],
        "变体/规格": config["variant"],
        "销售包数": pack_count,
        "采购数量": quantity,
        "采购单价RMB/包": purchase_cost,
        "订单运费RMB": config.get("order_shipping_rmb", 0),
        "订单优惠RMB": config.get("order_discount_rmb", 0),
        "包装长cm": package["length_cm"],
        "包装宽cm": package["width_cm"],
        "包装高cm": package["height_cm"],
        "包装重量g": package["weight_g"],
        "头程运费USD/包": config["first_leg_shipping_usd_per_bundle"],
        "目标最低利润率": config["target_margin_min"],
        "目标价格优势USD/每对比单位": config["target_price_gap_usd_per_comparison_unit"],
        "对比单位数量": config.get("comparison_unit_quantity", pack_count),
        "手动定价USD": manual_price,
        "竞品文件或链接": competitor_files,
        "备注": (
            f"组合套装：{component_summary(config)}。"
            f"采购单价={purchase_cost:.2f}元/套；可组成{quantity}套。"
            f"{config.get('notes', '')}"
        ),
    }
    source_notes = [
        ["项目", "内容"],
        ["销售单位", f"{config['variant']}；定价按一个买家实际购买的组合套装计算。"],
        ["采购单", f"{config['purchase_order']}；采购成本按三款各10片合并为一套。"],
        ["采购成本", f"{component_summary(config)}；合计 {purchase_cost:.2f} 元/套。"],
        ["采购数量", f"按最少可组成套数计算：{quantity} 套。"],
        [
            "尺寸重量",
            (
                f"整套最终包装：{package['length_cm']}*{package['width_cm']}*"
                f"{package['height_cm']}cm，{package['weight_g']}g。"
            ),
        ],
        ["竞品", f"已解析/录入 {len(competitors)} 个竞品，正式定价按{row['对比单位数量']}片对比单位折算。"],
        ["下一步", "确认草稿后运行 scripts/generate_final_pricing.py 生成正式定价结果。"],
    ]
    return [row], source_notes


def style_workbook(workbook):
    thin = Side(style="thin", color="D9D9D9")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 16


def write_workbook(config, rows, competitors, source_notes):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统一输入草稿"
    headers = [
        "产品批次", "SKU", "中文品名", "英文品名", "变体/规格", "销售包数", "采购数量",
        "采购单价RMB/包", "订单运费RMB", "订单优惠RMB", "包装长cm", "包装宽cm",
        "包装高cm", "包装重量g", "头程运费USD/包", "目标最低利润率",
        "目标价格优势USD/每对比单位", "对比单位数量", "手动定价USD", "竞品文件或链接", "备注",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    competitor_sheet = workbook.create_sheet("竞品解析")
    competitor_headers = ["文件", "ASIN", "页面主售价USD", "包数", "单件价USD", "容量ml", "标题"]
    competitor_sheet.append(competitor_headers)
    for item in competitors:
        competitor_sheet.append([item.get(header, "") for header in competitor_headers])

    source_sheet = workbook.create_sheet("来源与待确认")
    for row in source_notes:
        source_sheet.append(row)

    style_workbook(workbook)
    sheet.column_dimensions["T"].width = 56
    sheet.column_dimensions["U"].width = 90
    competitor_sheet.column_dimensions["A"].width = 42
    competitor_sheet.column_dimensions["G"].width = 80
    source_sheet.column_dimensions["B"].width = 120

    output_path = OUTPUT_DIR / f"{config['product_batch']}_统一输入草稿.xlsx"
    workbook.save(output_path)
    return output_path


def main():
    config = load_config()
    competitors = load_competitors(config)
    rows, source_notes = build_rows(config, competitors)
    output_path = write_workbook(config, rows, competitors, source_notes)
    print(output_path)


if __name__ == "__main__":
    main()
