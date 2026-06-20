#!/usr/bin/env python3
import json
import math
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "pricing_config.json"
OUTPUT_DIR = ROOT / "output"


FBA_FEE_TIERS = [
    (4 / 16, lambda price: price <= 3, 0.50, "4oz及以下且售价<=3"),
    (4 / 16, lambda price: price > 3, 0.88, "4oz及以下且售价>3"),
    (8 / 16, lambda price: True, 1.77, "4+~8oz"),
    (12 / 16, lambda price: True, 2.60, "8+~12oz"),
    (1.0, lambda price: True, 3.22, "12+~16oz"),
    (1.25, lambda price: True, 3.72, "1+~1.25lb"),
    (1.5, lambda price: True, 4.42, "1.25+~1.5lb"),
    (1.75, lambda price: True, 5.11, "1.5+~1.75lb"),
    (2.0, lambda price: True, 5.81, "1.75+~2lb"),
    (2.5, lambda price: True, 6.83, "2+~2.5lb"),
    (3.0, lambda price: True, 8.19, "2.5+~3lb"),
    (3.5, lambda price: True, 9.59, "3+~3.5lb"),
    (4.0, lambda price: True, 11.05, "3.5+~4lb"),
    (math.inf, lambda price: True, 12.51, "4lb+"),
]


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def to_float(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return default


def pick_latest_draft():
    drafts = [
        path for path in OUTPUT_DIR.glob("*_统一输入草稿.xlsx")
        if not path.name.startswith(("~$", ".~"))
    ]
    if not drafts:
        raise FileNotFoundError("没有找到 *_统一输入草稿.xlsx")
    return max(drafts, key=lambda path: path.stat().st_mtime)


def product_name_from_draft(path):
    return path.stem.replace("_统一输入草稿", "")


def fba_fee_for(weight_lb, price):
    for max_lb, predicate, fee, label in FBA_FEE_TIERS:
        if weight_lb <= max_lb and predicate(price):
            return fee, label
    return 12.51, "4lb+"


def metrics(row, price, total_qty, config):
    length = to_float(row["包装长cm"])
    width = to_float(row["包装宽cm"])
    height = to_float(row["包装高cm"])
    weight_g = to_float(row["包装重量g"])
    actual_lb = weight_g / 453.59237
    dimensional_lb = (
        (length / 2.54) * (width / 2.54) * (height / 2.54)
        / config["dimensional_weight_divisor"]
    )
    billable_lb = max(actual_lb, dimensional_lb)
    fba_fee, fba_tier = fba_fee_for(billable_lb, price)

    order_extra_rmb = (
        to_float(row.get("订单运费RMB"))
        + to_float(row.get("订单优惠RMB"))
    ) / total_qty
    unit_cost_rmb = to_float(row["采购单价RMB/包"]) + order_extra_rmb
    unit_cost_usd = unit_cost_rmb / config["currency_rate_rmb_to_usd"]
    first_leg = to_float(row.get("头程运费USD/包"), config["first_leg_shipping_usd"])
    referral = price * config["referral_fee_rate"]
    sold_profit = price - referral - fba_fee - first_leg - unit_cost_usd
    return_loss = (
        unit_cost_usd
        + referral * 0.2
        + first_leg
        + fba_fee
        + config["disposal_fee_usd"]
    )
    expected_profit = (
        (1 - config["return_rate"]) * sold_profit
        - config["return_rate"] * return_loss
    )
    margin = expected_profit / price if price else 0
    return {
        "actual_lb": actual_lb,
        "dimensional_lb": dimensional_lb,
        "billable_lb": billable_lb,
        "fba_fee": fba_fee,
        "fba_tier": fba_tier,
        "unit_cost_rmb": unit_cost_rmb,
        "unit_cost_usd": unit_cost_usd,
        "sold_profit": sold_profit,
        "return_loss": return_loss,
        "expected_profit": expected_profit,
        "margin": margin,
    }


def target_price(row, total_qty, config, target_margin):
    price = 2.99
    while price <= 39.99:
        if metrics(row, round(price, 2), total_qty, config)["margin"] >= target_margin:
            return round(price, 2)
        price += 0.01
    return None


def load_draft(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["统一输入草稿"]
    headers = [cell.value for cell in sheet[1]]
    rows = [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if values[0]
    ]
    competitor_sheet = workbook["竞品解析"]
    competitor_headers = [cell.value for cell in competitor_sheet[1]]
    competitors = [
        dict(zip(competitor_headers, values))
        for values in competitor_sheet.iter_rows(min_row=2, values_only=True)
        if values[0]
    ]
    return rows, competitors


def build_results(rows, competitors, config):
    comparison_unit = to_float(rows[0].get("对比单位数量"), 1)
    competitor_prices = [
        to_float(item.get("单件价USD")) * comparison_unit
        for item in competitors
        if to_float(item.get("单件价USD")) > 0
    ]
    competitor_min = min(competitor_prices) if competitor_prices else None
    total_qty = sum(to_float(row.get("采购数量")) for row in rows)
    results = []
    for row in rows:
        target_margin = to_float(row.get("目标最低利润率"), config["target_margin_min"])
        target_gap = to_float(row.get("目标价格优势USD/每对比单位"), config["target_price_gap_usd"])
        pack_count = max(to_float(row.get("销售包数"), 1), 1)
        competitive_price = (
            round((competitor_min - target_gap) / comparison_unit * pack_count, 2)
            if competitor_min is not None
            else None
        )
        margin_price = target_price(row, total_qty, config, target_margin)
        manual_price = to_float(row.get("手动定价USD"))

        if manual_price:
            recommended = manual_price
            conclusion = "手动定价"
        elif competitive_price is not None and metrics(row, competitive_price, total_qty, config)["margin"] >= target_margin:
            recommended = competitive_price
            conclusion = "可低价竞争"
        else:
            recommended = margin_price
            conclusion = "价格优势不足"

        recommended_metrics = metrics(row, recommended, total_qty, config) if recommended else None
        scenarios = []
        for label, price in [
            ("竞品优势价", competitive_price),
            ("竞品最低价", competitor_min / comparison_unit * pack_count if competitor_min else None),
            ("目标利润率价", margin_price),
            ("手动定价", manual_price if manual_price else None),
        ]:
            if price:
                scenarios.append((label, price, metrics(row, price, total_qty, config)))
        results.append({
            "row": row,
            "recommended": recommended,
            "metrics": recommended_metrics,
            "conclusion": conclusion,
            "scenarios": scenarios,
            "competitor_min": competitor_min,
            "comparison_unit": comparison_unit,
        })
    return results


def write_outputs(product_name, results, competitors):
    xlsx_path = OUTPUT_DIR / f"{product_name}_正式定价结果.xlsx"
    md_path = OUTPUT_DIR / f"{product_name}_正式定价报告.md"
    upload_path = OUTPUT_DIR / f"{product_name}_上品系统导入.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "建议售价"
    sheet.append([
        "SKU", "规格", "建议售价USD", "结论", "竞品最低对比价USD", "相对竞品差价",
        "计费重量lb", "FBA档位", "FBA费", "单件成本RMB", "单件成本USD",
        "期望利润", "期望利润率", "风险提示",
    ])
    for item in results:
        row = item["row"]
        metric = item["metrics"]
        gap = item["recommended"] - item["competitor_min"] if item["competitor_min"] else ""
        risk = "尺寸由部件尺寸推断，请确认最终包装尺寸" if "推断" in clean(row.get("备注")) else ""
        sheet.append([
            row["SKU"], row["变体/规格"], item["recommended"], item["conclusion"],
            item["competitor_min"], gap, metric["billable_lb"], metric["fba_tier"],
            metric["fba_fee"], metric["unit_cost_rmb"], metric["unit_cost_usd"],
            metric["expected_profit"], metric["margin"], risk,
        ])

    detail = workbook.create_sheet("价格档位")
    detail.append([
        "SKU", "规格", "档位", "售价USD", "计费重量lb", "FBA费",
        "售出利润", "退货亏损", "期望利润", "期望利润率",
    ])
    for item in results:
        row = item["row"]
        for label, price, metric in item["scenarios"]:
            detail.append([
                row["SKU"], row["变体/规格"], label, price, metric["billable_lb"],
                metric["fba_fee"], metric["sold_profit"], metric["return_loss"],
                metric["expected_profit"], metric["margin"],
            ])

    competitor_sheet = workbook.create_sheet("竞品")
    competitor_sheet.append(["文件", "ASIN", "页面主售价USD", "包数", "单件价USD", "容量ml", "标题"])
    for item in competitors:
        competitor_sheet.append([
            item.get("文件"), item.get("ASIN"), item.get("页面主售价USD"),
            item.get("包数"), item.get("单件价USD"), item.get("容量ml"),
            item.get("标题"),
        ])

    style_workbook(workbook)
    sheet.column_dimensions["N"].width = 34
    competitor_sheet.column_dimensions["G"].width = 80
    workbook.save(xlsx_path)
    write_upload_workbook(upload_path, results)
    write_markdown(md_path, product_name, results, competitors)
    return xlsx_path, md_path, upload_path


def write_upload_workbook(path, results):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "上品系统导入"
    sheet.append(["SKU", "规格", "销售包数", "包装尺寸cm", "包装重量g", "定价USD"])
    for item in results:
        row = item["row"]
        size = (
            f"{to_float(row.get('包装长cm')):g}*"
            f"{to_float(row.get('包装宽cm')):g}*"
            f"{to_float(row.get('包装高cm')):g}"
        )
        pack_count = to_float(row.get("销售包数"), 1)
        pack_label = f"{pack_count:g}双/包" if "手套" in clean(row.get("产品批次")) else f"{pack_count:g}/包"
        sheet.append([
            row.get("SKU"),
            row.get("变体/规格"),
            pack_label,
            size,
            f"{to_float(row.get('包装重量g')):g}g",
            item["recommended"],
        ])
    style_workbook(workbook)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 12
    for cell in sheet["F"]:
        if isinstance(cell.value, float):
            cell.number_format = "0.00"
    workbook.save(path)


def write_markdown(path, product_name, results, competitors):
    first = results[0]
    lines = [
        f"# {product_name}正式定价报告",
        "",
        f"- 竞品最低对比价：${first['competitor_min']:.2f}" if first["competitor_min"] else "- 竞品最低对比价：未识别",
        f"- 对比单位数量：{first['comparison_unit']:g}",
        "",
        "## 建议售价",
        "",
        "| SKU | 规格 | 建议售价 | 结论 | 计费重量 | FBA费 | 期望利润率 |",
        "| --- | --- | ---: | --- | ---: | ---: | ---: |",
    ]
    for item in results:
        row = item["row"]
        metric = item["metrics"]
        lines.append(
            f"| {row['SKU']} | {row['变体/规格']} | ${item['recommended']:.2f} | "
            f"{item['conclusion']} | {metric['billable_lb']:.2f} lb | "
            f"${metric['fba_fee']:.2f} | {metric['margin']:.1%} |"
        )
    lines.extend(["", "## 竞品", ""])
    for item in competitors:
        price = to_float(item.get("单件价USD"))
        lines.append(f"- ${price:.2f} / {item.get('ASIN')} / {item.get('标题')}")
    path.write_text("\n".join(lines), encoding="utf-8")


def style_workbook(workbook):
    thin = Side(style="thin", color="D9D9D9")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 16
    for sheet_name in ["建议售价", "价格档位"]:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"
    if "建议售价" in workbook.sheetnames:
        for cell in workbook["建议售价"]["M"]:
            if isinstance(cell.value, float):
                cell.number_format = "0.0%"
    if "价格档位" in workbook.sheetnames:
        for cell in workbook["价格档位"]["J"]:
            if isinstance(cell.value, float):
                cell.number_format = "0.0%"


def main():
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    draft_path = pick_latest_draft()
    product_name = product_name_from_draft(draft_path)
    rows, competitors = load_draft(draft_path)
    results = build_results(rows, competitors, config)
    xlsx_path, md_path, upload_path = write_outputs(product_name, results, competitors)
    print(xlsx_path)
    print(md_path)
    print(upload_path)


if __name__ == "__main__":
    main()
